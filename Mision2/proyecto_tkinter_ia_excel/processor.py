# processor.py
# Lógica de negocio: operaciones sobre Excel
import re
from openpyxl import load_workbook


def ejecutar_accion(instruccion, path):
    # Abre el archivo de ejemplo
    wb = load_workbook(path)
    ws = wb.active

    if instruccion["action"] == "clean_id":
        col = instruccion["column"]
        for fila in range(2, ws.max_row + 1):
            ws[f"H"+ str(fila) ]=''.join(filter(str.isdigit, str(ws[f"A{fila}"].value)))
            ws[f"{col}{fila}"] = ''.join(filter(str.isdigit, str(ws[f"{col}{fila}"].value)))

    elif instruccion["action"] == "merge_name":
        for fila in range(2, ws.max_row + 1):
            nombre = ws["B" + str(fila)].value or ""
            apellido = ws["C" + str(fila)].value or ""
            ws[f"H" + str(fila)] = f"{nombre} {apellido}".strip()

    wb.save(path)

def process_excel_safe(path):
    try:
        ejecutar_accion(path)
        return True, "Archivo procesado correctamente"
    except PermissionError:
        return( 
            False, "El archivo Excel esta abierto.\n"
            "por favor, cierrelo e intente nuevamente"
        )
    except KeyError:
        return False, "Hoja 'Datos' no encontrada"
    except Exception as e:
        return False, f"Error inesperado:{str(e)}"