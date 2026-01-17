import re
from openpyxl import load_workbook
#=====================================
# 😎  ANA MARIA RODRIGUEZ
# clean_id
#Elimina caracteres no numericos de un documento
#=====================================

def clean_id(value):
    if value is None:
        return ""
    return re.sub(r'\D','',str(value))

#=====================================
# 😎  ANA MARIA RODRIGUEZ
#funcion merge_name
#Une nombre y apellido en un solo campo
#=====================================
def merge_name(name, lastname):
    if name is None:
        name=""
    if lastname is None:
        lastname=""
    return f"{name} {lastname}".strip()

#=====================================
# 😎  ANA MARIA RODRIGUEZ
#excel cargar y procesar
#=====================================

def process_excel(path):
    wb=load_workbook(path)
    #Acceso a la hoja llamada "Datos"
    ws = wb["Datos"]
    #USAMOS LOS COMANDOS PARA QUE NOS SAQUE LOS COMANDOS CON UN COMANDO
    #recorrer todas las filas desde la fila 2
    for row in range(2,ws.max_row+1):
        ws[f"D{row}" ]=clean_id(ws[f"A{row}"].value)
    #columna E Nombre completo
        ws[f"E{row}"]=merge_name(
        ws[f"B{row}"].value, 
        ws[f"C{row}"].value
        )
        wb.save(path) 
def process_excel_safe(path):
    try:
        process_excel(path)
        return True, "Archivo procesado correctamente"
    except PermissionError:
        return( 
            False, "El archivo Excel esta abierto.\n"
            "por favor, cierrelo e intente nuevamente"
        )
    except KeyError:
        return False, "Hoja 'Datos' no encontrada"
    except Exception as e:
        return False, f"Error inesperado:{str(e)}"